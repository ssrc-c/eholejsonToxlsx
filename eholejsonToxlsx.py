import pandas as pd
import json, argparse, os
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment
from colorama import init,Fore
init(autoreset=True)

version = "1.0"

def logo():
    logo0 = r'''
    
                               
  ______ _____________   ____  
 /  ___//  ___/\_  __ \_/ ___\ 
 \___ \ \___ \  |  | \/\  \___ 
/____  >____  > |__|    \___  >
     \/     \/              \/ 
                                    Version {}
                                                By  山山而川'
                                                
                                                            
'''
    colored_logo = logo0.format(version)
    colored_logo = colored_logo.replace("____", Fore.YELLOW + "____" + Fore.RESET)

    print(colored_logo)

def usage():
    print('''
        用法:
            ehole3.0json转xlsx：    python CheckAlive.py -f ehole.json -o result
        参数：
            -f  --file     echole.json
            -o  --output   result文件名(无需带后缀)  ''')

def get_parser():
    parser = argparse.ArgumentParser(usage='python eholejsonToxlsx.py -f ehole.json -o result',
                                     description='ehole3.0输出的json文件转xlsx文件',
                                     )
    p = parser.add_argument_group('参数')
    p.add_argument("-f", "--file", type=str, help="echole输出的json文件")
    p.add_argument("-o", "--output", type=str, help="输出的文件名，无需带后缀")
    args = parser.parse_args()
    return args

json_data = []

def toExcel(jsonfile, out):
    if os.path.exists("%s.xlsx" %out):
        print("%s.xlsx文件已存在" %out)
        return
    print(Fore.RED+"[info]开始将%s转换为%s.xlsx ..." %(jsonfile, out))
    # 读取JSON文件逐行解析
    with open(jsonfile, 'r', encoding='utf-8') as f:
        for line in f:
            try:
                json_obj = json.loads(line)
                json_data.append(json_obj)
            except json.JSONDecodeError:
                continue

    # 创建DataFrame对象
    df = pd.DataFrame(json_data)

    # 转换为空值为""
    df = df.fillna("")

    # 处理"cms"列的值
    df['cms'] = df['cms'].apply(lambda x: x[0] if isinstance(x, list) else x)

    # 按照"cms"列进行降序排序
    df = df.sort_values(by="cms", ascending=False)

    # 重新排列列顺序
    columns = ["url", "cms", "server", "statuscode", "length", "title"]
    df = df[columns]

    # 将DataFrame保存为Excel文件
    df.to_excel(out + ".xlsx", index=False)

    # 打开Excel文件并填充JSON对应的内容

    # 加载Excel文件
    workbook = load_workbook(out + ".xlsx")

    # 选择第一个工作表
    worksheet = workbook.worksheets[0]

    # 设置列宽度
    worksheet.column_dimensions[get_column_letter(1)].width = 33
    worksheet.column_dimensions[get_column_letter(2)].width = 26
    worksheet.column_dimensions[get_column_letter(3)].width = 26
    worksheet.column_dimensions[get_column_letter(6)].width = 35

    # 设置单元格对齐方式
    alignment = Alignment(horizontal="center", vertical="center")
    for col_index, column in enumerate(columns, start=1):
        cell = worksheet.cell(row=1, column=col_index)
        cell.alignment = alignment

    # 遍历JSON数据，并填充到对应的单元格中
    for row_index, row_data in enumerate(df.to_dict(orient="records"), start=2):  # 从第二行开始，跳过标题行
        for col_index, key in enumerate(columns, start=1):
            value = str(row_data.get(key, ""))  # 将值转换为字符串
            value = value.replace("['","").replace("']","")
            worksheet.cell(row=row_index, column=col_index, value=value)

    # 固定首行
    worksheet.freeze_panes = 'A2'

    # 保存修改后的Excel文件
    workbook.save(out + ".xlsx")
    print(Fore.GREEN+"[info]转换完成，结果文件：%s.xlsx" %out)

def main():
    logo()
    args = get_parser()
    if args.file and args.output:
        toExcel(args.file, args.output)
    else:
        usage()   #如果没有输入任何参数则调用usage()
    
if __name__ == '__main__':
    main()
